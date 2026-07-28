"""Helper penulisan sheet Excel agar tampilannya mengikuti printout aplikasi.

Aturan tampilan disalin dari `FrontEnd/src/style.css` (bagian "Laporan / Cetak"):
nomor form rata kanan, judul tebal di tengah, tabel ber-border penuh dengan
header abu-abu, plus blok tanda tangan di kanan bawah. Ukuran kolom di sini
memakai satuan lebar Excel (± karakter), bukan piksel CSS.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Calibri"
FONT_SIZE = 9

_THIN = Side(style="thin", color="000000")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Warna mengikuti style cetak: header tabel abu-abu, blok hijau/kuning Form 2.b.
FILL_HEAD = PatternFill("solid", fgColor="E2E8F0")
FILL_LABEL = PatternFill("solid", fgColor="F8FAFC")
FILL_HIJAU = PatternFill("solid", fgColor="E8F5E9")
FILL_KUNING = PatternFill("solid", fgColor="FFF9C4")

TOP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
TOP_CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
MID_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def numbered(items: list[str]) -> str:
    """Daftar bernomor dalam satu sel (padanan <ol> pada cetakan HTML)."""
    return "\n".join(f"{i}. {t}" for i, t in enumerate(items, start=1))


def as_lines(value) -> str:
    """Field multi-nilai Form 2.a disimpan sebagai teks dipisah newline."""
    if not value:
        return ""
    return "\n".join(x for x in str(value).split("\n") if x)


class Sheet:
    """Penulis satu sheet: menjaga posisi baris & menerapkan gaya cetak."""

    def __init__(self, ws, widths: list[float], landscape: bool = True):
        self.ws = ws
        self.ncols = len(widths)
        self.row = 1
        self._head_rows: tuple[int, int] | None = None
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.page_setup.orientation = "landscape" if landscape else "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = ws.page_margins.right = 0.4
        ws.page_margins.top = ws.page_margins.bottom = 0.5
        # Fit-to-width: sepadan dengan @page landscape pada cetakan browser.
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

    # ----------------------------- primitif ---------------------------------
    def _cell(self, col: int, value, *, bold=False, align=None, fill=None,
              border=True, size=FONT_SIZE, italic=False, underline=None):
        c = self.ws.cell(row=self.row, column=col)
        c.value = value
        c.font = Font(name=FONT_NAME, size=size, bold=bold, italic=italic,
                      underline=underline)
        c.alignment = align or TOP_LEFT
        if border:
            c.border = BORDER
        if fill:
            c.fill = fill
        return c

    def _merge(self, col1: int, col2: int, row1: int | None = None,
               row2: int | None = None):
        if col2 <= col1 and (row2 or row1) == (row1 or self.row):
            return
        self.ws.merge_cells(
            start_row=row1 or self.row, start_column=col1,
            end_row=row2 or row1 or self.row, end_column=col2,
        )

    def skip(self, n: int = 1):
        self.row += n

    # ------------------------------ kop -------------------------------------
    def formno(self, text: str):
        """Nomor form rata kanan (mis. 'Lampiran 5' + 'Form 2.b')."""
        for line in text.split("\n"):
            self._cell(1, line, align=Alignment(horizontal="right"), border=False)
            self._merge(1, self.ncols)
            self.row += 1

    def title(self, text: str):
        self._cell(1, text.upper(), bold=True, size=FONT_SIZE + 2,
                   align=MID_CENTER, border=False)
        self._merge(1, self.ncols)
        self.row += 1

    def subtitle(self, text: str):
        self._cell(1, text, align=MID_CENTER, size=FONT_SIZE + 1, border=False)
        self._merge(1, self.ncols)
        self.row += 1

    # ----------------------------- tabel ------------------------------------
    def header(self, spec: list):
        """Header tabel. Item = str (kolom tunggal) atau (label, [sub, ...]).

        Bila ada grup, header memakai dua baris: kolom tunggal di-merge vertikal
        dan grup di-merge horizontal — sama seperti rowspan/colspan pada cetakan.
        """
        two = any(isinstance(s, tuple) for s in spec)
        r1 = self.row
        r2 = r1 + 1 if two else r1
        col = 1
        for item in spec:
            if isinstance(item, tuple):
                label, subs = item
                self._cell(col, label, bold=True, align=MID_CENTER, fill=FILL_HEAD)
                self._merge(col, col + len(subs) - 1, r1, r1)
                for j, sub in enumerate(subs):
                    c = self.ws.cell(row=r2, column=col + j)
                    c.value = sub
                    c.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
                    c.alignment = MID_CENTER
                    c.border = BORDER
                    c.fill = FILL_HEAD
                col += len(subs)
            else:
                self._cell(col, item, bold=True, align=MID_CENTER, fill=FILL_HEAD)
                if two:
                    self._merge(col, col, r1, r2)
                col += 1
        if two:
            # Sel bawah kolom tunggal tetap perlu border/fill walau ter-merge.
            for c in range(1, self.ncols + 1):
                cell = self.ws.cell(row=r2, column=c)
                if cell.border.left.style is None:
                    cell.border = BORDER
                    cell.fill = FILL_HEAD
        self.ws.row_dimensions[r1].height = 26
        if two:
            self.ws.row_dimensions[r2].height = 26
        self.row = r2 + 1
        self._head_rows = (r1, r2)
        # Ulangi baris header di tiap halaman cetak.
        self.ws.print_title_rows = f"{r1}:{r2}"

    def data(self, values: list, *, center: set[int] = frozenset(), bold=False,
             fills: dict[int, PatternFill] | None = None):
        """Satu baris data; `center` berisi indeks kolom (1-based) rata tengah."""
        for i, v in enumerate(values, start=1):
            self._cell(
                i, v if v not in (None, "") else None, bold=bold,
                align=TOP_CENTER if i in center else TOP_LEFT,
                fill=(fills or {}).get(i),
            )
        self.row += 1

    def group_row(self, label: str, *, first_col_blank=True):
        """Baris judul kelompok (mis. jenis risiko) yang di-merge selebar tabel."""
        start = 2 if first_col_blank else 1
        if first_col_blank:
            self._cell(1, None)
        self._cell(start, label, bold=True)
        for c in range(start + 1, self.ncols + 1):
            self._cell(c, None)
        self._merge(start, self.ncols)
        self.row += 1

    def empty_row(self, text: str = "Belum ada data"):
        self._cell(1, text, align=TOP_CENTER)
        for c in range(2, self.ncols + 1):
            self._cell(c, None)
        self._merge(1, self.ncols)
        self.row += 1

    def label_value(self, label: str, value, *, fill: PatternFill | None = None,
                    value_col: int = 2):
        """Baris label-isi (Form 2.a/2.b/2.c) — label kolom 1, isi di-merge."""
        self._cell(1, label, bold=True, fill=FILL_LABEL)
        self._cell(value_col, value if value not in (None, "") else "-", fill=fill)
        for c in range(value_col + 1, self.ncols + 1):
            self._cell(c, None, fill=fill)
        self._merge(value_col, self.ncols)
        self.row += 1

    # --------------------------- tanda tangan --------------------------------
    def sign(self, lines: list[str], nama: str, nip: str | None = None):
        """Blok tanda tangan rata kanan seperti .report-sign pada cetakan."""
        self.row += 1
        col = max(1, self.ncols - 2)
        for line in lines:
            self._cell(col, line, align=MID_CENTER, border=False)
            self._merge(col, self.ncols)
            self.row += 1
        self.row += 3  # ruang tanda tangan
        self._cell(col, nama, bold=True, align=MID_CENTER, border=False,
                   underline="single")
        self._merge(col, self.ncols)
        self.row += 1
        if nip:
            self._cell(col, f"NIP. {nip}", align=MID_CENTER, border=False)
            self._merge(col, self.ncols)
            self.row += 1
