"""Ekspor laporan ke Excel — satu sheet per form, mengikuti tampilan printout.

Sumber data sengaja memanggil fungsi endpoint yang sama dengan yang dipakai
halaman Cetak Laporan (`FrontEnd/src/views/Laporan.vue`), sehingga isi Excel
selalu sama dengan yang tampil/dicetak di layar. Susunan kolom & judul tiap
sheet mengikuti komponen di `FrontEnd/src/components/laporan/`.
"""
from io import BytesIO

from openpyxl import Workbook
from sqlmodel import Session, select

from app.models import (
    Infokom,
    KonteksOperasionalOpd,
    KonteksStrategisOpd,
    MonitoringPi,
    RtpCee,
)
from app.routers import cee, konteks, master, monitoring, risiko
from app.services.excel_sheet import (
    FILL_HIJAU,
    FILL_KUNING,
    Sheet,
    as_lines,
    numbered,
)

# Urutan & label bagian = `sections` pada Laporan.vue.
SECTIONS: list[tuple[str, str, str]] = [
    ("f1a", "Form 1.a", "Rekap Kuesioner CEE (Persepsi)"),
    ("f1b", "Form 1.b", "Reviu Dokumen CEE"),
    ("f1c", "Form 1.c", "Simpulan CEE"),
    ("f6", "Form 6", "RTP atas CEE"),
    ("f2a", "Form 2.a", "Konteks Strategis Pemda"),
    ("f2b", "Form 2.b", "Konteks Strategis OPD"),
    ("f2c", "Form 2.c", "Konteks Operasional OPD"),
    ("f3a", "Form 3.a", "Identifikasi Risiko Strategis Pemda"),
    ("f3b", "Form 3.b", "Identifikasi Risiko Strategis OPD"),
    ("f3c", "Form 3.c", "Identifikasi Risiko Operasional OPD"),
    ("f4", "Form 4", "Analisis Risiko"),
    ("f5", "Form 5", "Daftar Risiko Prioritas"),
    ("f7", "Form 7", "RTP atas Risiko"),
    ("f8", "Form 8", "Informasi & Komunikasi"),
    ("f9", "Form 9", "Rencana Pemantauan PI"),
    ("f10", "Form 10", "Monitoring Risk Event & RTP"),
]
SECTION_KEYS = [s[0] for s in SECTIONS]

JENIS_LABEL = {
    "strategis_pemda": "Risiko Strategis Pemda (Form 3.a)",
    "strategis_opd": "Risiko Strategis OPD (Form 3.b)",
    "operasional_opd": "Risiko Operasional OPD (Form 3.c)",
}
JENIS_ORDER = list(JENIS_LABEL)


class Ctx:
    """Data & identitas yang dipakai bersama seluruh sheet (dimuat sekali)."""

    def __init__(self, session: Session, opd_id: int, tahun: int):
        self.s = session
        self.opd_id = opd_id
        self.tahun = tahun
        self.opdctx = master.opd_print_context(opd_id, session) or {}
        self.opd_name = self.opdctx.get("nama_pd") or "-"

    def _list(self, model):
        return self.s.exec(
            select(model).where(model.opd_id == self.opd_id, model.tahun == self.tahun)
        ).all()

    def head(self, sheet: Sheet, formno: str, judul: str, subtitle: bool = True):
        sheet.formno(formno)
        sheet.title(judul)
        if subtitle:
            sheet.subtitle(f"{self.opd_name} — Tahun {self.tahun}")
        sheet.skip()

    def sign(self, sheet: Sheet, tahun_teks: str | None = None):
        o = self.opdctx
        sheet.sign(
            [
                tahun_teks or f"Kota Tegal, Desember {int(self.tahun) - 1}",
                o.get("jabatan_kepala") or f"Kepala {self.opd_name}",
            ],
            o.get("nama_kepala") or "(………………………………)",
            o.get("nip_kepala"),
        )


def _kelemahan(row: dict, fallback_key: str) -> str:
    """Sel 'Uraian Kelemahan': daftar bernomor bila ada, jika tidak teks bebas."""
    items = [k.get("uraian") for k in (row.get("kelemahan") or []) if k.get("uraian")]
    return numbered(items) if items else (row.get(fallback_key) or "")


# ------------------------------- CEE (Form 1 & 6) ----------------------------
def build_f1a(ctx: Ctx, wb: Workbook):
    data = cee.form_1a(ctx.opd_id, ctx.tahun, ctx.s)
    responden = data.get("responden") or []
    nresp = len(responden) or 1
    widths = [5, 55] + [5] * nresp + [8, 16]
    sh = Sheet(wb.create_sheet("Form 1.a"), widths)
    ctx.head(sh, "Form 1.a",
             "Rekapitulasi Hasil Kuesioner Penilaian Lingkungan Pengendalian (CEE)")
    kode = [r.get("kode_responden") or "" for r in responden] or ["—"]
    sh.header(["No", "Pertanyaan / Sub-unsur", ("Jawaban Responden", kode),
               "Modus", "Simpulan"])
    ncol = len(widths)
    for kat in data.get("kategori") or []:
        # Baris kategori: kode | nama (merge s/d kolom Modus) | simpulan.
        sh._cell(1, kat.get("kode"), bold=True, align=None)
        sh._cell(2, kat.get("nama"), bold=True)
        for c in range(3, ncol):
            sh._cell(c, None)
        sh._cell(ncol, kat.get("simpulan"))
        sh._merge(2, ncol - 1)
        sh.row += 1
        for i, pq in enumerate(kat.get("pertanyaan") or [], start=1):
            jawaban = pq.get("jawaban") or {}
            vals = [i, pq.get("pertanyaan")]
            vals += [jawaban.get(r["id"], jawaban.get(str(r["id"]))) for r in responden]
            if not responden:
                vals.append(None)
            vals += [pq.get("modus"), pq.get("simpulan")]
            sh.data(vals, center={1, *range(3, ncol + 1)})


def build_f1b(ctx: Ctx, wb: Workbook):
    rows = cee.form_1b(ctx.opd_id, ctx.tahun, ctx.s) or []
    sh = Sheet(wb.create_sheet("Form 1.b"), [6, 50, 14, 24, 50])
    ctx.head(sh, "Form 1.b", "Simpulan CEE Berdasarkan Reviu Dokumen")
    sh.header(["No", "Sub-unsur / Aspek", "Hasil Reviu", "Sumber Data",
               "Uraian Kelemahan"])
    for r in rows:
        sh.data([r.get("nomor"), r.get("aspek"), r.get("simpulan"),
                 r.get("sumber_data"), _kelemahan(r, "keterangan")],
                center={1, 3})
    if not rows:
        sh.empty_row()


def build_f1c(ctx: Ctx, wb: Workbook):
    rows = cee.form_1c(ctx.opd_id, ctx.tahun, ctx.s) or []
    sh = Sheet(wb.create_sheet("Form 1.c"), [6, 45, 14, 14, 14, 50])
    ctx.head(sh, "Form 1.c",
             "Simpulan Survei Persepsi atas Lingkungan Pengendalian Intern")
    sh.header(["No", "Sub-unsur", "Hasil Persepsi", "Hasil Dokumen", "Simpulan",
               "Penjelasan"])
    for r in rows:
        sh.data([r.get("no"), r.get("sub_unsur"), r.get("hasil_persepsi"),
                 r.get("hasil_dokumen"), r.get("simpulan"),
                 _kelemahan(r, "uraian_dokumen")], center={1, 3, 4, 5})
    if not rows:
        sh.empty_row()


def build_f6(ctx: Ctx, wb: Workbook):
    rows = ctx.s.exec(
        select(RtpCee)
        .where(RtpCee.opd_id == ctx.opd_id, RtpCee.tahun == ctx.tahun)
        .order_by(RtpCee.no_urut)
    ).all()
    sh = Sheet(wb.create_sheet("Form 6"), [6, 42, 42, 24, 13, 13])
    ctx.head(sh, "Form 6", "Penilaian Kegiatan Pengendalian (RTP atas CEE)")
    sh.header(["No", "Kondisi Kurang Memadai", "Rencana Tindak Pengendalian",
               "Penanggung Jawab", "Target", "Realisasi"])
    groups: dict[str, list] = {}
    for r in rows:
        groups.setdefault(r.aspek_cee or "(Tanpa Sub-unsur)", []).append(r)
    for aspek, items in groups.items():
        sh.group_row(aspek)
        for i, r in enumerate(items, start=1):
            sh.data([r.no_urut or i, r.kondisi_kerentanan, r.rencana_tindak,
                     r.pemilik_penanggung_jawab, r.target_waktu,
                     r.realisasi_waktu], center={1})
    if not rows:
        sh.empty_row()


# ------------------------------ Konteks (Form 2) -----------------------------
def build_f2a(ctx: Ctx, wb: Workbook):
    obj = konteks.get_pemda(ctx.tahun, ctx.s)
    d = obj.model_dump() if obj else {}
    sh = Sheet(wb.create_sheet("Form 2.a"), [42, 90], landscape=False)
    sh.formno("Form 2.a")
    sh.title("Penetapan Konteks Risiko Strategis Pemda")
    sh.subtitle(f"Tahun {ctx.tahun}")
    sh.skip()
    fields = [
        ("Periode yang Dinilai", "periode_dinilai", False),
        ("Visi", "visi", False),
        ("Misi Strategis RPJMD", "misi_strategis", True),
        ("Penetapan Konteks Tujuan Risiko Strategis Pemda",
         "penetapan_konteks_tujuan", True),
        ("Penetapan Konteks Sasaran Risiko Strategis Pemda",
         "penetapan_konteks_sasaran", True),
        ("Penetapan Konteks IKU Risiko Strategis Pemda",
         "penetapan_konteks_iku", True),
        ("Prioritas Pembangunan Daerah", "prioritas_pembangunan", True),
        ("Program Prioritas", "prioritas_program", True),
    ]
    for label, key, multi in fields:
        raw = d.get(key)
        value = numbered([x for x in as_lines(raw).split("\n") if x]) if multi \
            else (raw or "")
        sh.label_value(label, value)


def build_f2b(ctx: Ctx, wb: Workbook):
    rows = ctx.s.exec(
        select(KonteksStrategisOpd).where(
            KonteksStrategisOpd.opd_id == ctx.opd_id,
            KonteksStrategisOpd.tahun == ctx.tahun,
        ).order_by(KonteksStrategisOpd.id)
    ).all()
    tujuan = master.list_renstra_tujuan(ctx.opd_id, ctx.tahun, ctx.s)
    sasaran = master.list_renstra_sasaran(ctx.opd_id, ctx.tahun, ctx.s)
    iku = master.list_renstra_iku(ctx.opd_id, ctx.s)
    program = master.list_renstra_program(ctx.opd_id, ctx.s)

    def uniq(vals):
        out = []
        for v in vals:
            if v and v not in out:
                out.append(v)
        return out

    sh = Sheet(wb.create_sheet("Form 2.b"), [40, 95], landscape=False)
    sh.formno("Lampiran 5\nForm 2.b")
    sh.title("Penetapan Konteks Risiko Strategis OPD")
    sh.skip()
    # Blok hijau = seluruh data renstra OPD; blok kuning = entri Form 2.b.
    sh.label_value("Nama Pemda", "Pemerintah Kota Tegal", fill=FILL_HIJAU)
    sh.label_value("Tahun Penilaian", ctx.tahun, fill=FILL_HIJAU)
    sh.label_value("Periode yang Dinilai",
                   "; ".join(uniq([r.periode_dinilai for r in rows])),
                   fill=FILL_HIJAU)
    sh.label_value("Bidang Urusan", ctx.opdctx.get("bidang_urusan"), fill=FILL_HIJAU)
    sh.label_value("OPD yang Dinilai", ctx.opd_name, fill=FILL_HIJAU)
    sh.label_value("Sumber Data", "; ".join(uniq([r.sumber_data for r in rows])),
                   fill=FILL_HIJAU)
    sh.label_value("Tujuan Strategis",
                   "\n".join(o["value"] for o in tujuan), fill=FILL_HIJAU)
    sh.label_value("Sasaran Strategis",
                   "\n".join(o["value"] for o in sasaran), fill=FILL_HIJAU)
    sh.label_value("IKU Renstra OPD",
                   numbered([o["value"] for o in iku]), fill=FILL_HIJAU)
    sh.label_value("Program",
                   numbered([o["value"] for o in program]), fill=FILL_HIJAU)

    en_t = uniq([r.tujuan_strategis for r in rows])
    en_s = uniq([r.sasaran_strategis for r in rows])
    en_i = uniq([r.iku_renstra for r in rows])
    en_p = uniq([r.program for r in rows])
    blok = "\n".join([
        "Tujuan Strategis :", "\n".join(en_t) or "-",
        "Sasaran Strategis :", "\n".join(f"- {s}" for s in en_s) or "-",
        "IKU Renstra OPD :", numbered(en_i) or "-",
        "Program :", numbered(en_p) or "-",
    ])
    sh.label_value(
        "Tujuan, Sasaran, IKU dan Program yang akan dilakukan penilaian risiko",
        blok, fill=FILL_KUNING,
    )
    ctx.sign(sh)


def build_f2c(ctx: Ctx, wb: Workbook):
    renja = master.list_renja_subkegiatan(ctx.opd_id, ctx.tahun, ctx.s) or {}
    rows = ctx.s.exec(
        select(KonteksOperasionalOpd).where(
            KonteksOperasionalOpd.opd_id == ctx.opd_id,
            KonteksOperasionalOpd.tahun == ctx.tahun,
        )
    ).all()
    tujuan = master.list_renstra_tujuan(ctx.opd_id, ctx.tahun, ctx.s)
    checked_sub = {str(r.ref_subkegiatan) for r in rows if r.ref_subkegiatan}
    checked_ind = {str(r.ref_indikator) for r in rows if r.ref_indikator}

    sh = Sheet(wb.create_sheet("Form 2.c"), [6, 46, 46, 16, 12])
    sh.formno("Lampiran 6\nForm 2.c")
    sh.title("Penetapan Konteks Risiko Operasional OPD")
    sh.skip()
    sh.label_value("Nama Pemda", "Pemerintah Kota Tegal")
    sh.label_value("Tahun Penilaian", ctx.tahun)
    sh.label_value("Periode yang Dinilai",
                   f"DPA/APBD {ctx.opd_name} Tahun {ctx.tahun}")
    sh.label_value("Urusan Pemerintahan", ctx.opdctx.get("bidang_urusan"))
    sh.label_value("OPD yang Dinilai", ctx.opd_name)
    sh.label_value("Sumber Data", f"Renja {ctx.opd_name} Tahun {ctx.tahun}")
    sh.label_value("Tujuan Strategis", "\n".join(o["value"] for o in tujuan))
    sh.skip()

    sh.header(["No", "Program / Kegiatan / Sub Kegiatan",
               "Indikator Program / Kegiatan / Sub Kegiatan", "Target",
               "Dilakukan Penilaian Risiko"])
    # Ratakan hierarki Program -> Kegiatan -> Sub Kegiatan (lihat Form2c.vue).
    prog: dict[str, dict] = {}

    def ensure_prog(pid, nama):
        return prog.setdefault(str(pid), {"nm": nama, "ind": [], "keg": {}})

    def ensure_keg(g, kid, nama):
        return g["keg"].setdefault(str(kid), {"nm": nama, "ind": [], "subs": []})

    for s in renja.get("subkegiatan") or []:
        ensure_keg(ensure_prog(s["idprogram"], s["nm_program"]),
                   s["idkegiatan"], s["nm_kegiatan"])["subs"].append(s)
    for pi in renja.get("indikator_program") or []:
        ensure_prog(pi["idprogram"], pi["nm_program"])["ind"].append(pi)
    for ki in renja.get("indikator_kegiatan") or []:
        ensure_keg(ensure_prog(ki["idprogram"], ki["nm_program"]),
                   ki["idkegiatan"], ki["nm_kegiatan"])["ind"].append(ki)

    no = 0

    def emit(nama, indikator, target, sel, level):
        nonlocal no
        no += 1
        # Indentasi meniru padding-left lvl1/lvl2/lvl3 pada cetakan.
        prefix = {1: "", 2: "    ", 3: "        "}[level]
        sh.data([no, f"{prefix}{nama}" if nama else "", indikator, target,
                 "✓" if sel else ""], center={1, 5}, bold=level < 3)

    for g in prog.values():
        if g["ind"]:
            for i, pi in enumerate(g["ind"]):
                emit(g["nm"] if i == 0 else "", pi.get("tolok"), pi.get("target"),
                     str(pi.get("id")) in checked_ind, 1)
        else:
            emit(g["nm"], "", "", False, 1)
        for k in g["keg"].values():
            if k["ind"]:
                for i, ki in enumerate(k["ind"]):
                    emit(k["nm"] if i == 0 else "", ki.get("tolok"),
                         ki.get("target"), str(ki.get("id")) in checked_ind, 2)
            else:
                emit(k["nm"], "", "", False, 2)
            for s in k["subs"]:
                inds = s.get("indikator_list") or [{"tolok": "", "target": ""}]
                nama = f"{s.get('kode_sub_kegiatan') or ''} {s.get('nm_sub_kegiatan') or ''}".strip()
                for i, ind in enumerate(inds):
                    emit(nama if i == 0 else "", ind.get("tolok"),
                         ind.get("target"),
                         i == 0 and str(s["idsubkegiatan"]) in checked_sub, 3)
    if not no:
        sh.empty_row("Belum ada data renja")
    ctx.sign(sh)


# --------------------------- Identifikasi (Form 3) ---------------------------
F3_META = {
    "strategis_pemda": ("Form 3.a", "Identifikasi Risiko Strategis Pemerintah Daerah",
                        "Tujuan/Sasaran Strategis/Program", "Indikator Kinerja"),
    "strategis_opd": ("Form 3.b", "Identifikasi Risiko Strategis OPD",
                      "Tujuan/Sasaran Strategis", "Indikator Kinerja"),
    "operasional_opd": ("Form 3.c", "Identifikasi Risiko Operasional OPD",
                        "Kegiatan", "Indikator Keluaran"),
}


def _build_f3(ctx: Ctx, wb: Workbook, jenis: str):
    groups = risiko.form_4(ctx.opd_id, ctx.tahun, ctx.s)
    items = next((g["items"] for g in groups if g["jenis"] == jenis), [])
    formno, judul, kol_konteks, kol_indikator = F3_META[jenis]
    operasional = jenis == "operasional_opd"
    risiko_subs = (["Tahap", "Uraian", "Kode Risiko", "Pemilik"] if operasional
                   else ["Uraian", "Kode Risiko", "Pemilik"])
    widths = [5, 34, 26] + ([12] if operasional else []) + [34, 12, 16, 30, 11, 8, 30, 16]
    sh = Sheet(wb.create_sheet(formno), widths)
    ctx.head(sh, formno, f"Kertas Kerja {judul}")
    sh.header([
        "No", kol_konteks, kol_indikator,
        ("Risiko", risiko_subs),
        ("Sebab", ["Uraian", "Sumber"]),
        "C/UC",
        ("Dampak", ["Uraian", "Pihak yang Terkena"]),
    ])
    cuc_col = len(widths) - 2
    for i, it in enumerate(items, start=1):
        vals = [
            it.get("no_urut") or i,
            it.get("kegiatan") if operasional else it.get("tujuan_sasaran"),
            it.get("indikator_keluaran") if operasional else it.get("indikator_kinerja"),
        ]
        if operasional:
            vals.append(it.get("tahap_kegiatan"))
        vals += [it.get("uraian_risiko"), it.get("kode_risiko"),
                 it.get("pemilik_risiko"), it.get("sebab_uraian"),
                 it.get("sebab_sumber"), it.get("cuc"), it.get("dampak_uraian"),
                 it.get("dampak_pihak_terkena")]
        sh.data(vals, center={1, cuc_col})
    if not items:
        sh.empty_row()


def build_f3a(ctx, wb):
    _build_f3(ctx, wb, "strategis_pemda")


def build_f3b(ctx, wb):
    _build_f3(ctx, wb, "strategis_opd")


def build_f3c(ctx, wb):
    _build_f3(ctx, wb, "operasional_opd")


# ------------------------- Analisis & prioritas (4/5/7) ----------------------
def build_f4(ctx: Ctx, wb: Workbook):
    groups = risiko.form_4(ctx.opd_id, ctx.tahun, ctx.s)
    sh = Sheet(wb.create_sheet("Form 4"), [5, 60, 12, 10, 13, 9, 16])
    ctx.head(sh, "Form 4", "Kertas Kerja Hasil Analisis Risiko")
    sh.header(["No", "Risiko", "Kode", "Dampak", "Kemungkinan", "Skala", "Level"])
    total = 0
    for g in groups:
        sh.group_row(JENIS_LABEL.get(g["jenis"], g["jenis"]))
        for i, it in enumerate(g["items"], start=1):
            a = it.get("analisis") or {}
            total += 1
            sh.data([i, it.get("uraian_risiko"), it.get("kode_risiko"),
                     a.get("skala_dampak"), a.get("skala_kemungkinan"),
                     a.get("skala_risiko"), a.get("level")],
                    center={1, 4, 5, 6, 7})
    if not total:
        sh.empty_row()


def build_f5(ctx: Ctx, wb: Workbook):
    groups = risiko.form_5(ctx.opd_id, ctx.tahun, ctx.s)
    sh = Sheet(wb.create_sheet("Form 5"), [5, 46, 12, 9, 22, 36, 36])
    ctx.head(sh, "Form 5", "Kertas Kerja Daftar Risiko Prioritas")
    sh.header(["No", "Risiko Prioritas", "Kode", "Skala", "Pemilik", "Penyebab",
               "Dampak"])
    total = 0
    for g in groups:
        if not g["items"]:
            continue
        sh.group_row(JENIS_LABEL.get(g["jenis"], g["jenis"]))
        for i, it in enumerate(g["items"], start=1):
            total += 1
            sh.data([i, it.get("uraian_risiko"), it.get("kode_risiko"),
                     (it.get("analisis") or {}).get("skala_risiko"),
                     it.get("pemilik_risiko"), it.get("sebab_uraian"),
                     it.get("dampak_uraian")], center={1, 4})
    if not total:
        sh.empty_row("Belum ada risiko prioritas")


def build_f7(ctx: Ctx, wb: Workbook):
    rows = risiko.list_rtp(ctx.opd_id, ctx.tahun, ctx.s) or []
    sh = Sheet(wb.create_sheet("Form 7"), [5, 40, 12, 20, 34, 30, 40, 13])
    ctx.head(sh, "Form 7",
             "Penilaian Kegiatan Pengendalian (RTP atas Hasil Identifikasi Risiko)")
    sh.header(["No", "Risiko Prioritas", "Kode", "Pemilik",
               "Pengendalian yang Ada", "Celah", "Rencana Tindak Pengendalian",
               "Target"])
    for i, it in enumerate(rows, start=1):
        rtp = it.get("rtp") or {}
        sh.data([i, it.get("uraian_risiko"), it.get("kode_risiko"),
                 it.get("pemilik_risiko"), rtp.get("pengendalian_ada"),
                 rtp.get("celah_pengendalian"), rtp.get("rencana_tindak"),
                 rtp.get("target_waktu")], center={1})
    if not rows:
        sh.empty_row("Belum ada risiko prioritas")


# ---------------------- Komunikasi & pemantauan (8/9/10) ---------------------
def build_f8(ctx: Ctx, wb: Workbook):
    rows = ctx.s.exec(
        select(Infokom)
        .where(Infokom.opd_id == ctx.opd_id, Infokom.tahun == ctx.tahun)
        .order_by(Infokom.no_urut)
    ).all()
    sh = Sheet(wb.create_sheet("Form 8"), [5, 44, 26, 24, 24, 13, 13])
    ctx.head(sh, "Form 8", "Rencana & Realisasi Pengkomunikasian Pengendalian")
    sh.header(["No", "Kegiatan Pengendalian", "Media/Bentuk", "Penyedia",
               "Penerima", "Rencana", "Realisasi"])
    for i, r in enumerate(rows, start=1):
        sh.data([r.no_urut or i, r.kegiatan_pengendalian, r.media_bentuk,
                 r.penyedia_informasi, r.penerima_informasi, r.rencana_waktu,
                 r.realisasi_waktu], center={1})
    if not rows:
        sh.empty_row()


def _group_by_jenis(rows, key="jenis_risiko"):
    """Kelompokkan menurut urutan tetap 3.a -> 3.b -> 3.c, sisanya 'Lainnya'."""
    buckets: dict[str, list] = {}
    for r in rows:
        j = r.get(key) if isinstance(r, dict) else getattr(r, key, None)
        buckets.setdefault(j if j in JENIS_ORDER else "_lain", []).append(r)
    out = [(JENIS_LABEL[j], buckets[j]) for j in JENIS_ORDER if buckets.get(j)]
    if buckets.get("_lain"):
        out.append(("Lainnya", buckets["_lain"]))
    return out


def build_f9(ctx: Ctx, wb: Workbook):
    rows = ctx.s.exec(
        select(MonitoringPi)
        .where(MonitoringPi.opd_id == ctx.opd_id, MonitoringPi.tahun == ctx.tahun)
        .order_by(MonitoringPi.no_urut)
    ).all()
    sh = Sheet(wb.create_sheet("Form 9"), [5, 46, 34, 26, 13, 13])
    ctx.head(sh, "Form 9",
             "Rencana & Realisasi Pemantauan atas Kegiatan Pengendalian")
    sh.header(["No", "Kegiatan Pengendalian", "Metode Pemantauan",
               "Penanggung Jawab", "Rencana", "Realisasi"])
    groups = _group_by_jenis(rows)
    for label, items in groups:
        sh.group_row(label)
        for i, r in enumerate(items, start=1):
            sh.data([r.no_urut or i, r.kegiatan_pengendalian, r.metode_pemantauan,
                     r.penanggung_jawab, r.rencana_waktu, r.realisasi_waktu],
                    center={1})
    if not groups:
        sh.empty_row()


def build_f10(ctx: Ctx, wb: Workbook):
    rows = [r for r in monitoring.list_event(ctx.opd_id, ctx.tahun, ctx.s)
            if r.get("events")]
    sh = Sheet(wb.create_sheet("Form 10"), [12, 40, 12, 32, 32, 34, 30])
    ctx.head(sh, "Form 10",
             "Pencatatan Kejadian Risiko (Risk Event) & Pelaksanaan RTP")
    sh.header(["Kode", "Risiko", "Tanggal", "Sebab", "Dampak", "RTP",
               "Realisasi RTP"])
    groups = _group_by_jenis(rows)
    for label, items in groups:
        sh.group_row(label, first_col_blank=False)
        for r in items:
            for ev in r["events"]:
                sh.data([r.get("kode_risiko"), r.get("uraian_risiko"),
                         ev.get("tanggal_terjadi"), ev.get("sebab_kejadian"),
                         ev.get("dampak_kejadian"), ev.get("rtp"),
                         ev.get("realisasi_pelaksanaan_rtp")], center={3})
    if not groups:
        sh.empty_row("Belum ada kejadian dicatat")
    ctx.sign(sh, f"Kota Tegal, {ctx.tahun}")


BUILDERS = {
    "f1a": build_f1a, "f1b": build_f1b, "f1c": build_f1c, "f6": build_f6,
    "f2a": build_f2a, "f2b": build_f2b, "f2c": build_f2c,
    "f3a": build_f3a, "f3b": build_f3b, "f3c": build_f3c,
    "f4": build_f4, "f5": build_f5, "f7": build_f7,
    "f8": build_f8, "f9": build_f9, "f10": build_f10,
}


def build_workbook(session: Session, opd_id: int, tahun: int,
                   keys: list[str] | None = None) -> BytesIO:
    """Susun workbook berisi sheet untuk `keys` (default: seluruh form)."""
    wanted = [k for k in SECTION_KEYS if not keys or k in keys]
    ctx = Ctx(session, opd_id, tahun)
    wb = Workbook()
    wb.remove(wb.active)  # sheet bawaan tidak dipakai
    for key in wanted:
        BUILDERS[key](ctx, wb)
    if not wb.sheetnames:
        wb.create_sheet("Kosong")
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
